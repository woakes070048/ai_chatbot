# Copyright (c) 2026, Sanjay Kumar and contributors
# For license information, please see license.txt
"""
Excel (XLSX) and CSV text extraction.

Converts spreadsheet data into a readable text format
suitable for LLM processing.
"""

from __future__ import annotations

import csv
import io


def extract_excel_text(file_bytes: bytes) -> str:
	"""Extract tabular data from an XLSX file as formatted text.

	Reads all sheets. For each sheet, outputs the sheet name followed
	by a tab-separated representation of headers and data rows.

	Args:
		file_bytes: Raw XLSX file content.

	Returns:
		Formatted text representation of all sheets.
	"""
	try:
		from openpyxl import load_workbook
	except ImportError:
		import frappe

		frappe.throw("openpyxl is required for Excel file extraction. Install it with: pip install openpyxl")

	wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
	parts = []

	for sheet_name in wb.sheetnames:
		ws = wb[sheet_name]
		rows = list(ws.iter_rows(values_only=True))
		if not rows:
			continue

		lines = [f"--- Sheet: {sheet_name} ---"]
		for row in rows:
			cell_values = [_format_cell(cell) for cell in row]
			lines.append("\t".join(cell_values))

		parts.append("\n".join(lines))

	wb.close()
	return "\n\n".join(parts)


def extract_csv_text(file_bytes: bytes) -> str:
	"""Extract tabular data from a CSV file as formatted text.

	Args:
		file_bytes: Raw CSV file content.

	Returns:
		Tab-separated text representation of the CSV data.
	"""
	text = file_bytes.decode("utf-8", errors="replace")
	reader = csv.reader(io.StringIO(text))
	lines = []

	for row in reader:
		lines.append("\t".join(row))

	return "\n".join(lines)


def _format_cell(value) -> str:
	"""Format a cell value as a string.

	Handles None, datetime, and numeric types gracefully.
	"""
	if value is None:
		return ""
	return str(value)
